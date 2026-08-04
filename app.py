import os
import re
from datetime import datetime, timezone, timedelta
from zoneinfo import ZoneInfo
from flask import Flask, render_template, request, jsonify, send_file, g
from werkzeug.utils import secure_filename
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font
import io
import psycopg2
from psycopg2.extras import RealDictCursor
import json

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# ---------- 数据库连接 ----------
def get_db_connection():
    database_url = os.environ.get('DATABASE_URL')
    if not database_url:
        raise RuntimeError("DATABASE_URL environment variable not set")
    return psycopg2.connect(database_url, sslmode='require')

def init_db():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('''
        CREATE TABLE IF NOT EXISTS manifests (
            recipient TEXT PRIMARY KEY,
            data JSONB NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS scan_records (
            id SERIAL PRIMARY KEY,
            recipient TEXT NOT NULL,
            barcode TEXT NOT NULL,
            scanner_name TEXT,
            scanned_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(recipient, barcode)
        )
    ''')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_scan_recipient ON scan_records (recipient)')
    conn.commit()
    cur.close()
    conn.close()

# ---------- 辅助函数 ----------
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extract_custom_base_code(code):
    if not isinstance(code, str):
        return code
    match = re.search(r'^(.*)U(\d{6})$', code)
    if match:
        return match.group(1)
    return code

def generate_serial_list(base_code, count):
    return [f"{base_code}U{str(i+1).zfill(6)}" for i in range(count)]

def parse_excel(file_path):
    from openpyxl import load_workbook
    wb = load_workbook(file_path, data_only=True)
    sheet = wb['按箱出库'] if '按箱出库' in wb.sheetnames else wb.active

    header = [cell.value for cell in sheet[1]]
    try:
        recipient_col_idx = header.index('Recipient/收件人') + 1
        fba_col_idx = header.index('FBA货件ID/FBAShipmentID') + 1
        custom_col_idx = header.index('Custom box barcode/自定义箱条码') + 1
    except ValueError as e:
        raise Exception(f"未找到必要的列: {e}")

    data_rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue
        recipient = row[recipient_col_idx - 1] if len(row) >= recipient_col_idx else None
        fba_id = row[fba_col_idx - 1] if len(row) >= fba_col_idx else None
        custom_barcode = row[custom_col_idx - 1] if len(row) >= custom_col_idx else None
        data_rows.append({
            'recipient': recipient,
            'fba_id': fba_id,
            'custom_barcode': custom_barcode,
        })

    recipients = set()
    for row in data_rows:
        if row['recipient']:
            recipients.add(str(row['recipient']).strip())
    recipients = list(recipients)

    result = {}
    for rec in recipients:
        sub_rows = [r for r in data_rows if r['recipient'] and str(r['recipient']).strip() == rec]
        count_map = defaultdict(int)
        type_map = {}

        for row in sub_rows:
            # 原逻辑：优先 FBA，若 FBA 为空则使用 Custom
            code = row['fba_id'] if row['fba_id'] and str(row['fba_id']).strip() else row['custom_barcode']
            if not code or not str(code).strip():
                continue

            code_str = str(code).strip()
            if row['fba_id'] and str(row['fba_id']).strip():
                master = code_str
                type_tag = 'FBA'
            else:
                master = extract_custom_base_code(code_str)
                type_tag = 'Custom'

            count_map[master] += 1
            type_map[master] = type_tag

        items = []
        for master, count in count_map.items():
            serials = generate_serial_list(master, count)
            items.append({
                'master_code': master,
                'type': type_map.get(master, 'Unknown'),
                'total_boxes': count,
                'serials': serials
            })

        result[rec] = {'all_items': items}

    return result, recipients

# ---------- 数据库操作 ----------
def save_manifest(recipient, data):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('''
        INSERT INTO manifests (recipient, data, updated_at)
        VALUES (%s, %s, CURRENT_TIMESTAMP)
        ON CONFLICT (recipient) DO UPDATE
        SET data = EXCLUDED.data, updated_at = CURRENT_TIMESTAMP
    ''', (recipient, json.dumps(data)))
    conn.commit()
    cur.close()
    conn.close()

def load_manifest(recipient):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute('SELECT data FROM manifests WHERE recipient=%s', (recipient,))
    row = cur.fetchone()
    cur.close()
    conn.close()
    if row:
        return row['data']   # ← 直接返回 dict，因为 psycopg2 已自动转换
    return None

def load_all_recipients():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('SELECT recipient FROM manifests')
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [row[0] for row in rows]

# ---------- 路由 ----------
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    if not allowed_file(file.filename):
        return jsonify({'error': 'File type not allowed'}), 400

    filename = secure_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)

    try:
        data, recipients = parse_excel(filepath)
        for rec, rec_data in data.items():
            save_manifest(rec, rec_data)
        return jsonify({
            'success': True,
            'recipients': recipients,
            'data': data
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/get_data', methods=['GET'])
def get_data():
    recipients = load_all_recipients()
    if not recipients:
        return jsonify({'success': False, 'message': '暂无数据，请先上传出库表'}), 404
    data = {}
    for rec in recipients:
        manifest = load_manifest(rec)
        if manifest:
            data[rec] = manifest
    return jsonify({
        'success': True,
        'recipients': recipients,
        'data': data
    })

@app.route('/scan', methods=['POST'])
def scan_barcode():
    req = request.json
    recipient = req.get('recipient')
    barcode = req.get('barcode')
    scanner_name = req.get('scanner_name', '').strip()
    if not recipient or not barcode:
        return jsonify({'error': 'Missing parameters'}), 400
    if not scanner_name:
        return jsonify({'error': '请先输入扫描员姓名'}), 400

    manifest = load_manifest(recipient)
    if not manifest:
        return jsonify({'error': 'Recipient not found'}), 400

    all_items = manifest.get('all_items', [])
    valid_codes = set()
    for item in all_items:
        for code in item['serials']:
            valid_codes.add(code)
    if barcode not in valid_codes:
        return jsonify({'match': False, 'message': '无效条码'}), 200

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('SELECT 1 FROM scan_records WHERE recipient=%s AND barcode=%s', (recipient, barcode))
    if cur.fetchone():
        cur.close()
        conn.close()
        return jsonify({'match': True, 'already_scanned': True, 'message': '已扫过'}), 200

    cur.execute('INSERT INTO scan_records (recipient, barcode, scanner_name) VALUES (%s, %s, %s)',
                (recipient, barcode, scanner_name))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'match': True, 'already_scanned': False, 'message': '扫描成功'}), 200

@app.route('/status/<recipient>')
def get_status(recipient):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute('SELECT barcode, scanner_name, scanned_at FROM scan_records WHERE recipient=%s', (recipient,))
    rows = cur.fetchall()
    cur.close()
    conn.close()
    scanned = [{'barcode': row['barcode'], 'scanner': row['scanner_name'], 'time': row['scanned_at']} for row in rows]
    return jsonify({'scanned': scanned})

@app.route('/export', methods=['GET'])
def export_excel():
    recipient = request.args.get('recipient')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not recipient:
        return jsonify({'error': '缺少收件人参数'}), 400

    manifest = load_manifest(recipient)
    if not manifest:
        return jsonify({'error': '收件人数据不存在'}), 404

    items = manifest.get('all_items', [])
    if not items:
        return jsonify({'error': '无数据可导出'}), 400

    mexico_tz = ZoneInfo("America/Mexico_City")
    utc_start = None
    utc_end = None
    if start_date:
        local_dt = datetime.strptime(start_date, '%Y-%m-%d').replace(tzinfo=mexico_tz)
        utc_start = local_dt.astimezone(timezone.utc).replace(tzinfo=None)
    if end_date:
        local_dt = datetime.strptime(end_date, '%Y-%m-%d').replace(tzinfo=mexico_tz)
        next_day = local_dt + timedelta(days=1)
        utc_end = next_day.astimezone(timezone.utc).replace(tzinfo=None)

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    query = 'SELECT barcode, scanner_name, scanned_at FROM scan_records WHERE recipient=%s'
    params = [recipient]
    if utc_start:
        query += ' AND scanned_at >= %s'
        params.append(utc_start.strftime('%Y-%m-%d %H:%M:%S'))
    if utc_end:
        query += ' AND scanned_at < %s'
        params.append(utc_end.strftime('%Y-%m-%d %H:%M:%S'))
    cur.execute(query, params)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    scan_info = {}
    for row in rows:
        scan_info[row['barcode']] = {
            'scanner': row['scanner_name'],
            'time': row['scanned_at']
        }

    wb = Workbook()
    ws = wb.active
    ws.title = '复核清单'
    headers = ['主码Master', '类型Tipo', '总箱数Veces', '条码Codigo', '扫描状态Estado de escaneo',
               '箱数Caja', '日期Fecha', '时间Hora', '扫描员Escaneador']
    ws.append(headers)
    for col in range(1, len(headers)+1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    for item in items:
        for code in item['serials']:
            info = scan_info.get(code)
            if info:
                status = '已扫/Si'
                caja = 1
                utc_time = info['time']
                if isinstance(utc_time, str):
                    try:
                        utc_dt = datetime.strptime(utc_time, '%Y-%m-%d %H:%M:%S')
                    except ValueError:
                        utc_dt = datetime.fromisoformat(utc_time.replace('Z', '+00:00'))
                else:
                    utc_dt = utc_time
                if utc_dt.tzinfo is None:
                    utc_dt = utc_dt.replace(tzinfo=timezone.utc)
                mexico_dt = utc_dt.astimezone(mexico_tz)
                fecha = mexico_dt.strftime('%Y-%m-%d')
                hora = mexico_dt.strftime('%H:%M:%S')
                scanner = info['scanner'] or ''
            else:
                status = '未扫/No'
                caja = 0
                fecha = ''
                hora = ''
                scanner = ''

            ws.append([
                item['master_code'],
                item['type'],
                item['total_boxes'],
                code,
                status,
                caja,
                fecha,
                hora,
                scanner
            ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
                     as_attachment=True,
                     download_name=f'复核清单_{recipient}_{datetime.now().strftime("%Y%m%d")}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/reset/<recipient>')
def reset_recipient(recipient):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('DELETE FROM scan_records WHERE recipient=%s', (recipient,))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'success': True})

@app.route('/reset_master/<recipient>/<master_code>', methods=['POST'])
def reset_master(recipient, master_code):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('DELETE FROM scan_records WHERE recipient=%s AND barcode LIKE %s',
                (recipient, master_code + 'U%'))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'success': True})

if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
